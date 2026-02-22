"""Web UI routes blueprint."""

import os
import uuid
import tempfile

import pandas as pd
from io import BytesIO
from flask import Blueprint, request, jsonify, render_template, send_file
from datetime import datetime, timezone
from db import get_db
from config import PROCESSED_FILES, allowed_file, get_cf_user
from phone import validate_phone_numbers, convert_to_international, convert_to_local
from transliteration import is_hebrew
from providers import get_provider, get_all_providers
from lookup import lookup, translate_and_score
from app_logger import log_event
from input_validator import validate_file_size

web_bp = Blueprint("web", __name__)


@web_bp.route("/web/apis", methods=["GET"])
def web_apis():
    """Return list of available APIs."""
    apis = [
        {"name": p.display_name, "available": p.is_configured}
        for p in get_all_providers()
    ]
    return jsonify({"apis": apis})


@web_bp.route("/")
@web_bp.route("/web")
def web_index():
    """Serve the web interface."""
    return render_template("index.html")


@web_bp.route("/web/query")
def web_query_page():
    """Serve the single query interface."""
    return render_template("query.html")


@web_bp.route("/web/query", methods=["POST"])
def web_query():
    """Process single phone query."""
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "JSON body required"})

    phone = data.get("phone", "").strip()
    cal_name = data.get("name", "").strip()
    refresh_days = data.get("refresh_days", 7)
    apis_str = data.get("apis", "me")
    selected_apis = [a.strip().lower() for a in apis_str.split(',') if a.strip()]

    # Per-provider cache_only flags
    cache_only_flags = {}
    for api_name in selected_apis:
        cache_only_flags[api_name] = data.get(f"{api_name}_cache_only", False)

    if not phone:
        return jsonify({"success": False, "error": "מספר טלפון נדרש"})

    phone_list = convert_to_international([phone])
    phone = phone_list[0]

    if not validate_phone_numbers([phone]):
        return jsonify({"success": False, "error": "מספר טלפון לא תקין"})

    if cal_name and not is_hebrew(cal_name):
        return jsonify({"success": False, "error": "שם איש קשר חייב להיות בעברית"})

    try:
        db = get_db()
        result = {"phone_number": phone, "cal_name": cal_name}
        any_api_called = False
        log_kwargs = {}

        for api_name in selected_apis:
            provider = get_provider(api_name)
            if not provider or not provider.is_configured:
                continue

            try:
                provider_data, api_called, from_cache = lookup(
                    provider, db, phone, cal_name, refresh_days,
                    cache_only=cache_only_flags.get(api_name, False),
                )
                result.update(provider_data)
                result["phone_number"] = phone
                result["cal_name"] = cal_name

                if cal_name:
                    translate_and_score(provider, result, cal_name, db)

                if api_called:
                    any_api_called = True
                    primary_key = provider.get_primary_name_key()
                    log_kwargs[f"{api_name}_result"] = "success" if result.get(primary_key) else "fail"

                log_kwargs[f"{api_name}_api_call"] = api_called
                log_kwargs[f"{api_name}_cache"] = from_cache

            except Exception as provider_error:
                primary_key = provider.get_primary_name_key()
                result[primary_key] = f"ERROR: {provider_error}"
                result[f"{provider.name}.matching"] = 0

        log_event(
            user=get_cf_user() or "",
            action="query",
            phone=phone,
            me_api_call=log_kwargs.get("me_api_call", False),
            sync_api_call=log_kwargs.get("sync_api_call", False),
            me_cache=log_kwargs.get("me_cache", False),
            sync_cache=log_kwargs.get("sync_cache", False),
            me_result=log_kwargs.get("me_result", ""),
            sync_result=log_kwargs.get("sync_result", ""),
            datetime_str=datetime.now(timezone.utc).isoformat(),
        )

        return jsonify({
            "success": True,
            "result": result,
            "from_cache": not any_api_called,
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@web_bp.route("/web/process", methods=["POST"])
def web_process():
    """Process uploaded file via web interface."""
    import base64

    # Accept both JSON (base64 file) and multipart form data
    if request.is_json:
        json_data = request.get_json()
        file_data_b64 = json_data.get("file_data")
        if not file_data_b64:
            return jsonify({"success": False, "error": "No file uploaded"})
        original_filename = json_data.get("filename", "upload.xlsx")
        if not allowed_file(original_filename):
            return jsonify({"success": False, "error": "Invalid file type. Only .xlsx and .csv files are allowed"}), 400
        file_bytes = base64.b64decode(file_data_b64)
        validate_file_size(len(file_bytes))
        file = BytesIO(file_bytes)
        try:
            refresh_days = int(json_data.get("refresh_days", 7))
        except (ValueError, TypeError):
            refresh_days = 7
        apis_str = json_data.get("apis", "me")
        selected_apis = [a.strip().lower() for a in apis_str.split(",") if a.strip()]
        if not selected_apis:
            selected_apis = ["me"]
        cache_only_flags = {}
        for api_name in selected_apis:
            cache_only_flags[api_name] = bool(json_data.get(f"{api_name}_cache_only", False))
    else:
        if 'file' not in request.files:
            return jsonify({"success": False, "error": "No file uploaded"})
        uploaded = request.files['file']
        if uploaded.filename == '':
            return jsonify({"success": False, "error": "No file selected"})
        if not allowed_file(uploaded.filename):
            return jsonify({"success": False, "error": "Invalid file type. Only .xlsx and .csv files are allowed"}), 400
        original_filename = request.form.get('original_filename') or uploaded.filename
        uploaded.seek(0, 2)
        file_size = uploaded.tell()
        uploaded.seek(0)
        validate_file_size(file_size)
        file = uploaded
        try:
            refresh_days = int(request.form.get('refresh_days', 7))
        except ValueError:
            refresh_days = 7
        apis_str = request.form.get('apis', 'me')
        selected_apis = [a.strip().lower() for a in apis_str.split(',') if a.strip()]
        if not selected_apis:
            selected_apis = ['me']
        cache_only_flags = {}
        for api_name in selected_apis:
            cache_only_flags[api_name] = request.form.get(f'{api_name}_cache_only', '').lower() == 'true'

    # Filter to configured providers only
    active_providers = []
    for api_name in selected_apis:
        provider = get_provider(api_name)
        if provider and provider.is_configured:
            active_providers.append(provider)

    if not active_providers:
        return jsonify({"success": False, "error": "No configured API providers selected"})

    # Build file suffix from active providers
    file_suffix = "_" + "_".join(p.name for p in active_providers)

    try:
        # Read file
        filename = original_filename.lower()
        if filename.endswith('.csv'):
            data = pd.read_csv(file, dtype=str, header=None)
        elif filename.endswith('.xlsx'):
            data = pd.read_excel(file, dtype=str, header=None)
        else:
            return jsonify({"success": False, "error": "פורמט קובץ לא נתמך. השתמש ב-.xlsx או .csv"})

        if len(data) > 100000:
            return jsonify({"success": False, "error": f"File too large: {len(data)} rows (max 100,000)"}), 400

        if data.shape[1] < 2:
            return jsonify({"success": False, "error": f"הקובץ חייב להכיל לפחות 2 עמודות (טלפון, שם). נמצאו {data.shape[1]} עמודות"})

        if len(data) == 0:
            return jsonify({"success": False, "error": "הקובץ ריק"})

        # Detect and remove header row (check first 2 columns only)
        first_row = data.iloc[0]
        header_indicators = ['phone', 'טלפון', 'מספר', 'first', 'last', 'שם', 'name', 'פרטי', 'משפחה']
        is_header = False
        for cell in first_row.iloc[:2]:
            cell_str = str(cell).lower().strip() if pd.notna(cell) else ""
            if any(indicator in cell_str for indicator in header_indicators):
                is_header = True
                break
            if cell_str and not cell_str.replace('+', '').replace('-', '').replace(' ', '').isdigit():
                is_header = True
                break

        # Save or generate headers — first two columns always fixed
        if is_header:
            original_headers = [str(cell).strip() if pd.notna(cell) else "" for cell in first_row]
        else:
            original_headers = [""] * data.shape[1]
        original_headers[0] = "טלפון"
        original_headers[1] = "שם לקוח"
        start_row = 1 if is_header else 0
        data = data.iloc[start_row:].reset_index(drop=True)

        if len(data) == 0:
            return jsonify({"success": False, "error": "הקובץ ריק (רק שורת כותרת)"})

        # Validate rows
        errors = []
        valid_rows = []

        def clean_name(name):
            if not name:
                return ""
            return name.replace("'", "").replace("\u2019", "").replace("`", "").strip()

        def is_valid_phone(phone):
            if not phone:
                return False
            phone = str(phone).strip().replace('-', '').replace(' ', '').replace('+', '')
            if phone.startswith('05') or phone.startswith('07'):
                return len(phone) >= 9 and len(phone) <= 10 and phone.isdigit()
            if phone.startswith('5') or phone.startswith('7'):
                return len(phone) == 9 and phone.isdigit()
            if phone.startswith('972'):
                return len(phone) >= 11 and len(phone) <= 12 and phone.isdigit()
            return False

        for idx, row in data.iterrows():
            excel_row = idx + 2 if is_header else idx + 1
            row_errors = []

            phone = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            cal_name = clean_name(str(row.iloc[1])) if pd.notna(row.iloc[1]) else ""

            if not phone:
                row_errors.append(f"שורה {excel_row}, עמודה A: טלפון ריק")
            elif not is_valid_phone(phone):
                row_errors.append(f"שורה {excel_row}, עמודה A: טלפון לא תקין '{phone}'")

            if not cal_name:
                row_errors.append(f"שורה {excel_row}, עמודה B: שם ריק")
            elif not is_hebrew(cal_name):
                row_errors.append(f"שורה {excel_row}, עמודה B: שם חייב להיות בעברית '{cal_name}'")

            if row_errors:
                errors.extend(row_errors)
            else:
                valid_rows.append({
                    "phone": phone,
                    "cal_name": cal_name,
                })

        if errors:
            error_list = "\n".join(errors[:20])
            if len(errors) > 20:
                error_list += f"\n... ועוד {len(errors) - 20} שגיאות"
            return jsonify({"success": False, "error": f"שגיאות בקובץ:\n{error_list}"})

        if not valid_rows:
            return jsonify({"success": False, "error": "לא נמצאו שורות תקינות בקובץ"})

        # Convert phones to international format
        for row in valid_rows:
            converted = convert_to_international([row["phone"]])
            row["phone"] = converted[0]

        results = []
        cache_counts = {p.name: 0 for p in active_providers}
        api_counts = {p.name: 0 for p in active_providers}

        db = get_db()
        log_username = get_cf_user() or ""

        # Process each row
        for row_data in valid_rows:
            phone = row_data["phone"]
            cal_name = row_data["cal_name"]

            result = {"phone_number": phone, "cal_name": cal_name}
            row_log = {}

            try:
                for provider in active_providers:
                    pname = provider.name
                    provider_data, api_called, from_cache = lookup(
                        provider, db, phone, cal_name, refresh_days,
                        cache_only=cache_only_flags.get(pname, False),
                    )

                    if api_called:
                        api_counts[pname] += 1
                        primary_key = provider.get_primary_name_key()
                        row_log[f"{pname}_result"] = "success" if provider_data.get(primary_key) else "fail"
                    elif from_cache:
                        cache_counts[pname] += 1

                    result.update(provider_data)
                    result["phone_number"] = phone
                    result["cal_name"] = cal_name

                    # Source indicator
                    if api_called:
                        result[f"{pname}.source"] = "API"
                    elif from_cache:
                        result[f"{pname}.source"] = "cache"
                    else:
                        result[f"{pname}.source"] = "cache-only"

                    row_log[f"{pname}_api_call"] = api_called
                    row_log[f"{pname}_cache"] = from_cache

                results.append(result)

                log_event(
                    user=log_username,
                    action="process_file",
                    phone=phone,
                    filename=original_filename,
                    me_api_call=row_log.get("me_api_call", False),
                    sync_api_call=row_log.get("sync_api_call", False),
                    me_cache=row_log.get("me_cache", False),
                    sync_cache=row_log.get("sync_cache", False),
                    me_result=row_log.get("me_result", ""),
                    sync_result=row_log.get("sync_result", ""),
                    datetime_str=datetime.now(timezone.utc).isoformat(),
                )

            except Exception as e:
                for provider in active_providers:
                    primary_key = provider.get_primary_name_key()
                    result[primary_key] = f"ERROR: {e}"
                results.append(result)

        # Post-process: translations and matching scores
        for result in results:
            cal_name = str(result.get("cal_name", "") or "")
            for provider in active_providers:
                translate_and_score(provider, result, cal_name, db)

        # Build result columns
        result_df = pd.DataFrame(results).astype(str)

        result_columns = []
        for provider in active_providers:
            result_columns.extend(provider.excel_columns)
        result_cols_df = result_df[[c for c in result_columns if c in result_df.columns]]

        # Find first empty column (starting from col 2) to insert results
        insert_col = 2  # Default: right after phone + name
        for col_idx in range(2, data.shape[1]):
            col_data = data.iloc[:, col_idx]
            is_empty = col_data.isna().all() or (col_data.astype(str).str.strip().isin(['', 'nan', 'None'])).all()
            if is_empty:
                break
            insert_col = col_idx + 1

        # Convert phone column to local format (0XX...)
        data.iloc[:, 0] = data.iloc[:, 0].apply(lambda x: convert_to_local(x) if pd.notna(x) else x)

        # Build output: original columns up to insert point + results + remaining original columns
        out_df = data.iloc[:, :insert_col].copy()
        # Rename original columns using saved headers
        if original_headers:
            for col_idx in range(insert_col):
                if col_idx < len(original_headers) and original_headers[col_idx]:
                    out_df.rename(columns={col_idx: original_headers[col_idx]}, inplace=True)
        for col_name in result_cols_df.columns:
            out_df[col_name] = result_cols_df[col_name].values
        if insert_col < data.shape[1]:
            for col_idx in range(insert_col, data.shape[1]):
                orig_col = data.iloc[:, col_idx]
                if not (orig_col.isna().all() or (orig_col.astype(str).str.strip() == '').all()):
                    header = original_headers[col_idx] if original_headers and col_idx < len(original_headers) and original_headers[col_idx] else f"col_{col_idx}"
                    out_df[header] = orig_col.values

        file_id = str(uuid.uuid4())
        temp_path = os.path.join(tempfile.gettempdir(), f"result_{file_id}.xlsx")
        out_df.to_excel(temp_path, index=False, engine="openpyxl")

        # Style the Excel file
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        wb = load_workbook(temp_path)
        ws = wb.active

        # Force phone column (col A) to text format to preserve leading zero
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = numbers.FORMAT_TEXT
                cell.value = str(cell.value) if cell.value is not None else ""

        # Blue headers with white text
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Color matching score cells
        green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        green_font = Font(color="2E7D32", bold=True)
        yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        yellow_font = Font(color="F57F17", bold=True)
        red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        red_font = Font(color="C62828", bold=True)

        # Find special columns
        col_map = {cell.value: cell.column for cell in ws[1] if cell.value}
        score_cols = [col_map[c] for c in col_map if c.endswith(".matching")]
        tier_cols = [col_map[c] for c in col_map if c.endswith(".risk_tier")]
        translated_cols = [col_map[c] for c in col_map if c.endswith(".translated")]

        sand_fill = PatternFill(start_color="F5DEB3", end_color="F5DEB3", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in score_cols + tier_cols:
                cell = row[col_idx - 1]
                try:
                    val = int(float(str(cell.value or 0)))
                except (ValueError, TypeError):
                    val = -1
                if col_idx in tier_cols:
                    tier = str(cell.value or "").upper()
                    if tier == "HIGH":
                        cell.fill, cell.font = green_fill, green_font
                    elif tier == "MEDIUM":
                        cell.fill, cell.font = yellow_fill, yellow_font
                    elif tier in ("LOW", "VERY LOW"):
                        cell.fill, cell.font = red_fill, red_font
                else:
                    if val >= 70:
                        cell.fill, cell.font = green_fill, green_font
                    elif val >= 50:
                        cell.fill, cell.font = yellow_fill, yellow_font
                    elif val >= 0:
                        cell.fill, cell.font = red_fill, red_font
            for col_idx in translated_cols:
                cell = row[col_idx - 1]
                val = str(cell.value or "").strip()
                if val and val not in ("None", "nan", ""):
                    cell.fill = sand_fill

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            padding = 1 if col[0].column == 1 else 0
            ws.column_dimensions[col[0].column_letter].width = max_len + padding

        wb.save(temp_path)

        PROCESSED_FILES[file_id] = {
            "path": temp_path,
            "created": datetime.now(),
            "original_name": os.path.splitext(original_filename)[0] + file_suffix + ".xlsx"
        }

        total_from_cache = sum(cache_counts.values())
        total_api_calls = sum(api_counts.values())

        return jsonify({
            "success": True,
            "file_id": file_id,
            "total": len(results),
            "from_cache": total_from_cache,
            "api_calls": total_api_calls
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@web_bp.route("/web/download/<file_id>")
def web_download(file_id):
    """Download processed file. File is deleted after download."""
    if file_id not in PROCESSED_FILES:
        return jsonify({"error": "File not found"}), 404

    file_info = PROCESSED_FILES.pop(file_id)
    file_path = file_info["path"]

    if not os.path.exists(file_path):
        return jsonify({"error": "File expired"}), 404

    with open(file_path, "rb") as f:
        data = BytesIO(f.read())
    try:
        os.remove(file_path)
    except OSError:
        pass

    return send_file(
        data,
        as_attachment=True,
        download_name=file_info["original_name"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
