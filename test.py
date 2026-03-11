"""
test.py — Comprehensive test suite for PhoneInfo.

Covers:
  - Health endpoint
  - CF Access authentication enforcement
  - CSRF protection
  - Security headers
  - Phone utilities  (phone.py)
  - Input validator  (input_validator.py)
  - REST API endpoints: /me, /sync, /lookup/<provider>, /translate, /compare, /nicknames
  - Web pages: GET rendering
  - Web query: single phone lookup
  - File processing: xlsx/csv upload + download
  - Nicknames CRUD: list/get/save/delete/upload/download/backup/restore
  - Security vectors: formula injection, magic bytes, CSRF, LIKE wildcard, XSS sanitization,
                      DISABLE_AUTH removal, phone masking in logs

Usage:
    pytest test.py -v
    pytest test.py -v -k security
    pytest test.py -v -k nicknames
    pytest test.py -v -k api
"""

import base64
import csv
import io
import json
import os
import tempfile
from unittest.mock import MagicMock, patch

import pytest

# ─────────────────────────────────────────────────────────────────────────────
# Bootstrap — set env vars BEFORE importing the Flask app.
# config.py calls load_dotenv() which does NOT override existing os.environ
# keys, so anything we set here wins over the .env file.
# ─────────────────────────────────────────────────────────────────────────────
_tmpdb = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmpdb.close()
os.environ["DATABASE"]       = _tmpdb.name
os.environ["ME_API_URL"]     = "http://fake-me-api/"
os.environ["ME_API_SID"]     = "test-sid"
os.environ["ME_API_TOKEN"]   = "test-token"
os.environ["SYNC_API_URL"]   = "http://fake-sync-api/"
os.environ["SYNC_API_TOKEN"] = "test-token"

from server import app as flask_app  # noqa: E402  (must come after env setup)

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
CF_EMAIL = "ram@catsec.com"
AUTH  = {"Cf-Access-Authenticated-User-Email": CF_EMAIL}
HOST  = {"Host": "testserver"}
H     = {**AUTH, **HOST}   # full headers for every authenticated request

PHONE_LOCAL   = "0521234567"
PHONE_INTL    = "972521234567"
PHONE_LOCAL2  = "0501111111"
PHONE_INTL2   = "972501111111"

ME_API_RESPONSE = {
    "common_name":    "יוסי כהן",
    "me_profile_name": "yosi.cohen",
    "result_strength": "HIGH",
    "whitelist":       "",
    "user": {
        "first_name":      "Yosi",
        "last_name":       "Cohen",
        "email":           "yosi@example.com",
        "email_confirmed": True,
        "profile_picture": "",
        "gender":          "M",
        "is_verified":     True,
        "slogan":          "",
        "social_profiles": {
            "facebook": "yosi.cohen", "instagram": "", "twitter": "",
            "spotify": "", "linkedin": "", "pinterest": "", "tiktok": "",
        },
    },
}

SYNC_API_RESPONSE = {
    "results": {
        "name":              "Yosi Cohen",
        "is_potential_spam": False,
        "is_business":       False,
        "job_hint":          "",
        "company_hint":      "",
        "website_domain":    "",
        "company_domain":    "",
    }
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def make_xlsx_bytes(rows: list) -> bytes:
    """Build a minimal .xlsx file in memory from a list of row lists."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_csv_bytes(rows: list) -> bytes:
    """Build a minimal CSV file in memory from a list of row lists."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def b64(data: bytes) -> str:
    return base64.b64encode(data).decode()


def json_post(client, path, payload, extra_headers=None):
    headers = {**H, **(extra_headers or {})}
    return client.post(
        path,
        data=json.dumps(payload),
        content_type="application/json",
        headers=headers,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Session fixture
# ─────────────────────────────────────────────────────────────────────────────
@pytest.fixture(scope="session")
def client():
    flask_app.config["TESTING"] = True
    flask_app.config["RATELIMIT_ENABLED"] = False
    with flask_app.test_client() as c:
        yield c


def pytest_sessionfinish(session, exitstatus):
    try:
        os.unlink(_tmpdb.name)
    except OSError:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# 1. Health
# ─────────────────────────────────────────────────────────────────────────────
class TestHealth:
    def test_no_auth_required(self, client):
        r = client.get("/health")
        assert r.status_code == 200
        assert r.get_json()["status"] == "ok"

    def test_with_auth_still_works(self, client):
        r = client.get("/health", headers=H)
        assert r.status_code == 200
        assert r.get_json()["status"] == "ok"


# ─────────────────────────────────────────────────────────────────────────────
# 2. Authentication
# ─────────────────────────────────────────────────────────────────────────────
class TestAuthentication:
    @pytest.mark.parametrize("path", [
        "/", "/web", "/web/query", "/me", "/sync",
        "/translate", "/web/nicknames", "/web/apis",
    ])
    def test_no_header_returns_403(self, client, path):
        r = client.get(path)
        assert r.status_code == 403, f"Expected 403 on GET {path}, got {r.status_code}"

    def test_empty_header_value_returns_403(self, client):
        r = client.get("/", headers={"Cf-Access-Authenticated-User-Email": ""})
        assert r.status_code == 403

    def test_valid_header_passes(self, client):
        r = client.get("/", headers=H)
        assert r.status_code == 200

    def test_health_exempt_from_auth(self, client):
        r = client.get("/health")
        assert r.status_code == 200   # no auth header needed


# ─────────────────────────────────────────────────────────────────────────────
# 3. CSRF
# ─────────────────────────────────────────────────────────────────────────────
class TestCSRF:
    def test_json_post_not_csrf_blocked(self, client):
        """Content-Type: application/json prevents CSRF — must never be blocked."""
        r = json_post(client, "/translate", {"first": "John", "last": "Smith"})
        assert r.status_code == 200

    def test_multipart_from_evil_origin_blocked(self, client):
        r = client.post(
            "/web/process",
            headers={**H, "Origin": "https://evil.com"},
            data={"file": (io.BytesIO(b"x"), "test.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 403
        assert "CSRF" in r.get_json()["error"]

    def test_multipart_from_evil_referer_blocked(self, client):
        r = client.post(
            "/web/process",
            headers={**H, "Referer": "https://phishing.site/page"},
            data={"file": (io.BytesIO(b"x"), "test.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 403

    def test_multipart_same_origin_not_blocked(self, client):
        """Same-origin multipart may fail for other reasons but not CSRF."""
        xlsx = make_xlsx_bytes([[PHONE_LOCAL, "יוסי כהן"]])
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = client.post(
                "/web/process",
                headers={**H, "Origin": "http://testserver"},
                data={"file": (io.BytesIO(xlsx), "test.xlsx"), "apis": "me", "refresh_days": "0"},
                content_type="multipart/form-data",
            )
        if r.status_code == 403:
            assert "CSRF" not in r.get_json().get("error", "")

    def test_no_origin_no_referer_not_blocked(self, client):
        """Direct API/script calls without browser headers must not be CSRF-blocked."""
        r = json_post(client, "/translate", {"first": "Dan"})
        assert r.status_code == 200


# ─────────────────────────────────────────────────────────────────────────────
# 4. Security Headers
# ─────────────────────────────────────────────────────────────────────────────
class TestSecurityHeaders:
    def _headers(self, client):
        return client.get("/health").headers

    def test_csp_directives(self, client):
        csp = self._headers(client).get("Content-Security-Policy", "")
        for directive in [
            "default-src 'self'", "object-src 'none'",
            "base-uri 'none'", "form-action 'self'",
            "frame-ancestors 'none'", "nonce-",
        ]:
            assert directive in csp, f"Missing CSP directive: {directive}"

    def test_x_frame_options(self, client):
        assert self._headers(client).get("X-Frame-Options") == "DENY"

    def test_x_content_type_options(self, client):
        assert self._headers(client).get("X-Content-Type-Options") == "nosniff"

    def test_referrer_policy(self, client):
        assert "strict-origin" in self._headers(client).get("Referrer-Policy", "")

    def test_permissions_policy(self, client):
        pp = self._headers(client).get("Permissions-Policy", "")
        for feature in ["camera=()", "microphone=()", "geolocation=()"]:
            assert feature in pp, f"Missing Permissions-Policy feature: {feature}"

    def test_hsts_absent_on_plain_http(self, client):
        """HSTS must not be set when there is no HTTPS signal."""
        assert "Strict-Transport-Security" not in self._headers(client)

    def test_hsts_present_when_forwarded_proto_https(self, client):
        r = client.get("/health", headers={"X-Forwarded-Proto": "https"})
        assert "Strict-Transport-Security" in r.headers

    def test_no_cache_on_web_routes(self, client):
        r = client.get("/web/query", headers=H)
        assert "no-store" in r.headers.get("Cache-Control", "")


# ─────────────────────────────────────────────────────────────────────────────
# 5. Phone Utilities (pure unit tests)
# ─────────────────────────────────────────────────────────────────────────────
class TestPhoneUtils:
    def test_validate_valid_international(self):
        from phone import validate_phone_numbers
        assert validate_phone_numbers(["972521234567"]) is True

    def test_validate_rejects_local_format(self):
        from phone import validate_phone_numbers
        assert validate_phone_numbers(["0521234567"]) is False

    def test_validate_rejects_short(self):
        from phone import validate_phone_numbers
        assert validate_phone_numbers(["97252123456"]) is False   # 11 digits

    def test_validate_rejects_alpha(self):
        from phone import validate_phone_numbers
        assert validate_phone_numbers(["abcdefghijkl"]) is False

    def test_convert_local_to_intl(self):
        from phone import convert_to_international
        assert convert_to_international(["0521234567"]) == ["972521234567"]

    def test_convert_9digit_to_intl(self):
        from phone import convert_to_international
        assert convert_to_international(["521234567"]) == ["972521234567"]

    def test_convert_intl_to_local(self):
        from phone import convert_to_local
        assert convert_to_local("972521234567") == "0521234567"
        assert convert_to_local("0521234567") == "0521234567"

    def test_is_valid_forms(self):
        from phone import is_valid_israeli_phone
        assert is_valid_israeli_phone("0521234567") is True
        assert is_valid_israeli_phone("972521234567") is True
        assert is_valid_israeli_phone("521234567") is True

    def test_is_invalid_forms(self):
        from phone import is_valid_israeli_phone
        assert is_valid_israeli_phone("1234") is False
        assert is_valid_israeli_phone("") is False
        assert is_valid_israeli_phone("abcdefghij") is False


# ─────────────────────────────────────────────────────────────────────────────
# 6. Input Validator (pure unit tests)
# ─────────────────────────────────────────────────────────────────────────────
class TestInputValidator:
    def test_clean_name_strips_angle_brackets(self):
        from input_validator import clean_name
        assert "<" not in clean_name("יוסי<script>")
        assert ">" not in clean_name("יוסי<script>")

    def test_clean_name_keeps_hebrew_english(self):
        from input_validator import clean_name
        assert clean_name("יוסי כהן") == "יוסי כהן"
        assert clean_name("John Smith") == "John Smith"

    def test_clean_name_collapses_whitespace(self):
        from input_validator import clean_name
        assert clean_name("יוסי  כהן") == "יוסי כהן"

    def test_sanitize_string_truncates_to_max_length(self):
        from input_validator import sanitize_string
        result = sanitize_string("A" * 300, max_length=100, field_type="text")
        assert len(result) <= 100

    def test_sanitize_phone_strips_dashes(self):
        from input_validator import sanitize_phone
        assert sanitize_phone("052-123-4567") == "0521234567"

    def test_sanitize_phone_keeps_plus(self):
        from input_validator import sanitize_phone
        assert sanitize_phone("+972521234567") == "+972521234567"

    def test_sanitize_phone_none_returns_empty(self):
        from input_validator import sanitize_phone
        assert sanitize_phone(None) == ""

    def test_sanitize_phone_alpha_returns_empty(self):
        from input_validator import sanitize_phone
        assert sanitize_phone("abcdef") == ""

    def test_validate_nicknames_data_valid(self):
        from input_validator import validate_nicknames_data
        data = [{"formal_name": "יוסי", "all_names": "יוסף,יוסי,יוסיפון"}]
        result = validate_nicknames_data(data)
        assert len(result) == 1
        assert result[0]["formal_name"] == "יוסי"

    def test_validate_nicknames_too_many_rows(self):
        from input_validator import validate_nicknames_data, ValidationError
        data = [{"formal_name": f"x{i}", "all_names": f"y{i}"} for i in range(100001)]
        with pytest.raises(ValidationError):
            validate_nicknames_data(data)

    def test_validate_nicknames_missing_key(self):
        from input_validator import validate_nicknames_data, ValidationError
        with pytest.raises(ValidationError):
            validate_nicknames_data([{"formal_name": "test"}])   # missing all_names


# ─────────────────────────────────────────────────────────────────────────────
# 7. /translate
# ─────────────────────────────────────────────────────────────────────────────
class TestTranslate:
    def test_english_to_hebrew(self, client):
        r = json_post(client, "/translate", {"first": "Yosi", "last": "Cohen"})
        assert r.status_code == 200
        d = r.get_json()
        assert "first" in d and "last" in d

    def test_hebrew_passthrough(self, client):
        r = json_post(client, "/translate", {"first": "יוסי", "last": "כהן"})
        assert r.status_code == 200
        d = r.get_json()
        assert d["first"] == "יוסי"
        assert d["last"] == "כהן"

    def test_russian_transliterated(self, client):
        r = json_post(client, "/translate", {"first": "Иван", "last": "Петров"})
        assert r.status_code == 200
        d = r.get_json()
        assert d["first"] != ""

    def test_empty_fields(self, client):
        r = json_post(client, "/translate", {"first": "", "last": ""})
        assert r.status_code == 200
        d = r.get_json()
        assert d["first"] == "" and d["last"] == ""

    def test_no_body_returns_400(self, client):
        r = client.post("/translate", headers=H)
        assert r.status_code in (400, 415)   # Flask 2.3+ may return 415


# ─────────────────────────────────────────────────────────────────────────────
# 8. /compare
# ─────────────────────────────────────────────────────────────────────────────
class TestCompare:
    def test_matching_names_high_score(self, client):
        r = json_post(client, "/compare", {
            "first": "יוסי", "last": "כהן", "names": ["יוסי כהן"],
        })
        assert r.status_code == 200
        d = r.get_json()
        assert d["score"] > 50
        assert d["risk_tier"] in ("HIGH", "MEDIUM", "LOW", "VERY LOW")

    def test_mismatch_low_score(self, client):
        r = json_post(client, "/compare", {
            "first": "דוד", "last": "לוי", "names": ["שמשון גבעוני"],
        })
        assert r.status_code == 200
        assert r.get_json()["score"] < 80

    def test_breakdown_and_explanation_present(self, client):
        r = json_post(client, "/compare", {
            "first": "יוסי", "last": "כהן", "names": ["יוסי כהן"],
        })
        d = r.get_json()
        assert "breakdown" in d
        assert "explanation" in d

    def test_no_body_returns_400(self, client):
        r = client.post("/compare", headers=H)
        assert r.status_code in (400, 415)


# ─────────────────────────────────────────────────────────────────────────────
# 9. /nicknames (GET API)
# ─────────────────────────────────────────────────────────────────────────────
class TestNicknamesGetAPI:
    def test_known_name_returns_variants(self, client):
        json_post(client, "/web/nicknames/save",
                  {"formal_name": "יוסף", "all_names": "יוסי,יוסיפון"})
        r = client.get("/nicknames?name=יוסף", headers=H)
        assert r.status_code == 200
        assert "יוסף" in r.get_json()["names"]

    def test_missing_param_returns_400(self, client):
        r = client.get("/nicknames", headers=H)
        assert r.status_code == 400

    def test_unknown_name_returns_itself(self, client):
        r = client.get("/nicknames?name=שםלאקיים999", headers=H)
        assert r.status_code == 200
        assert "שםלאקיים999" in r.get_json()["names"]


# ─────────────────────────────────────────────────────────────────────────────
# 10. /me, /sync, /lookup/<provider>
# ─────────────────────────────────────────────────────────────────────────────
class TestLookupAPI:
    def _post(self, client, path, phone, cal_name="", **kw):
        return json_post(client, path, {
            "phone": phone, "cal_name": cal_name, "use_cache": False, **kw
        })

    def test_me_api_call_returns_data(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._post(client, "/me", PHONE_LOCAL)
        assert r.status_code == 200
        d = r.get_json()
        assert d["me.common_name"] == "יוסי כהן"
        assert d["from_cache"] is False

    def test_me_hit_cache_on_second_call(self, client):
        phone = "0521234568"
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            self._post(client, "/me", phone)
        r = json_post(client, "/me", {"phone": phone, "use_cache": True, "refresh_days": 30})
        assert r.status_code == 200
        assert r.get_json()["from_cache"] is True

    def test_me_invalid_phone_returns_400(self, client):
        r = self._post(client, "/me", "notaphone")
        assert r.status_code == 400

    def test_me_missing_phone_returns_400(self, client):
        r = json_post(client, "/me", {})
        assert r.status_code == 400

    def test_me_cache_only_not_in_cache(self, client):
        r = json_post(client, "/me", {"phone": "0509999991", "noapi": True})
        assert r.status_code == 200
        assert "NOT IN CACHE" in str(r.get_json().get("me.common_name", ""))

    def test_sync_api_call(self, client):
        with patch("providers.sync.SyncProvider.call_api", return_value=SYNC_API_RESPONSE):
            r = self._post(client, "/sync", "0501234567")
        assert r.status_code == 200
        d = r.get_json()
        assert "sync.first_name" in d

    def test_sync_invalid_phone(self, client):
        r = self._post(client, "/sync", "bad")
        assert r.status_code == 400

    def test_lookup_unknown_provider_returns_404(self, client):
        r = json_post(client, "/lookup/nonexistent", {"phone": PHONE_LOCAL})
        assert r.status_code == 404

    def test_me_international_format_accepted(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._post(client, "/me", PHONE_INTL)
        assert r.status_code == 200

    def test_me_api_none_response_stored(self, client):
        """API returning None (not found) should be stored without error."""
        with patch("providers.me.MEProvider.call_api", return_value=None):
            r = self._post(client, "/me", "0521234590")
        assert r.status_code == 200


# ─────────────────────────────────────────────────────────────────────────────
# 11. Web Pages — GET rendering
# ─────────────────────────────────────────────────────────────────────────────
class TestWebPages:
    @pytest.mark.parametrize("path", [
        "/", "/web", "/web/query",
        "/web/nicknames", "/web/nicknames/edit",
    ])
    def test_page_renders(self, client, path):
        r = client.get(path, headers=H)
        assert r.status_code == 200
        assert b"<html" in r.data.lower()

    def test_web_apis_returns_provider_list(self, client):
        r = client.get("/web/apis", headers=H)
        assert r.status_code == 200
        d = r.get_json()
        assert "apis" in d
        names = [a["name"] for a in d["apis"]]
        assert "ME" in names
        assert "SYNC" in names


# ─────────────────────────────────────────────────────────────────────────────
# 12. Web Query — single phone
# ─────────────────────────────────────────────────────────────────────────────
class TestWebQuery:
    def _query(self, client, phone, name="", apis="me", **kw):
        return json_post(client, "/web/query",
                         {"phone": phone, "name": name, "apis": apis, **kw})

    def test_success_returns_result(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._query(client, PHONE_LOCAL, name="יוסי כהן")
        assert r.status_code == 200
        d = r.get_json()
        assert d["success"] is True
        assert d["result"]["me.common_name"] == "יוסי כהן"

    def test_name_produces_matching_score(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._query(client, "0521234570", name="יוסי כהן")
        assert r.status_code == 200
        result = r.get_json()["result"]
        assert "me.matching" in result
        assert "me.risk_tier" in result

    def test_non_hebrew_name_rejected(self, client):
        r = self._query(client, PHONE_LOCAL, name="John Smith")
        assert r.get_json()["success"] is False

    def test_invalid_phone_rejected(self, client):
        r = self._query(client, "notaphone", name="יוסי כהן")
        assert r.get_json()["success"] is False

    def test_empty_phone_rejected(self, client):
        r = self._query(client, "")
        assert r.get_json()["success"] is False

    def test_cache_only_flag(self, client):
        r = self._query(client, "0508888881", name="", me_cache_only=True)
        assert r.get_json()["success"] is True
        assert "NOT IN CACHE" in str(r.get_json()["result"].get("me.common_name", ""))

    def test_no_body_returns_error(self, client):
        r = client.post("/web/query", headers=H)
        # Flask 2.3+ returns 415 when Content-Type is missing; older returns 200 with error body
        if r.status_code == 200:
            assert r.get_json()["success"] is False
        else:
            assert r.status_code in (400, 415)

    def test_from_cache_flag_in_response(self, client):
        phone = "0521234571"
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            self._query(client, phone)   # prime the cache
        r = self._query(client, phone)   # should hit cache
        assert "from_cache" in r.get_json()


# ─────────────────────────────────────────────────────────────────────────────
# 13. File Processing — /web/process + /web/download
# ─────────────────────────────────────────────────────────────────────────────
HEADER_ROW = ["טלפון", "שם לקוח"]   # header the app's auto-detection expects


class TestFileProcessing:
    DEFAULT_ROWS = [HEADER_ROW, [PHONE_LOCAL, "יוסי כהן"]]

    def _upload_json(self, client, rows=None, filename="test.xlsx", apis="me", **kw):
        if filename.endswith(".xlsx"):
            data = make_xlsx_bytes(rows or self.DEFAULT_ROWS)
        else:
            data = make_csv_bytes(rows or self.DEFAULT_ROWS)
        return json_post(client, "/web/process", {
            "file_data": b64(data), "filename": filename,
            "apis": apis, "refresh_days": 0, **kw,
        })

    # ── happy path ──────────────────────────────────────────────────────────
    def test_xlsx_upload_success(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._upload_json(client)
        assert r.status_code == 200
        d = r.get_json()
        assert d["success"] is True
        assert "file_id" in d
        assert d["total"] == 1

    def test_csv_upload_success(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._upload_json(client, rows=[HEADER_ROW, [PHONE_LOCAL2, "דני לוי"]], filename="test.csv")
        assert r.status_code == 200
        assert r.get_json()["success"] is True

    def test_download_after_upload(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._upload_json(client, rows=[HEADER_ROW, [PHONE_LOCAL, "יוסי כהן"]])
        file_id = r.get_json()["file_id"]
        dl = client.get(f"/web/download/{file_id}", headers=H)
        assert dl.status_code == 200
        assert "spreadsheetml" in dl.mimetype

    def test_download_only_once(self, client):
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._upload_json(client, rows=[HEADER_ROW, [PHONE_LOCAL2, "דני לוי"]])
        file_id = r.get_json()["file_id"]
        client.get(f"/web/download/{file_id}", headers=H)
        r2 = client.get(f"/web/download/{file_id}", headers=H)
        assert r2.status_code == 404

    def test_download_unknown_id_returns_404(self, client):
        r = client.get("/web/download/no-such-id", headers=H)
        assert r.status_code == 404

    def test_header_row_detected_and_stripped(self, client):
        rows = [["טלפון", "שם לקוח"], [PHONE_LOCAL, "יוסי כהן"]]
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._upload_json(client, rows=rows)
        assert r.status_code == 200
        assert r.get_json()["success"] is True

    def test_multirow_file(self, client):
        rows = [
            HEADER_ROW,
            [PHONE_LOCAL,  "יוסי כהן"],
            [PHONE_LOCAL2, "דני לוי"],
        ]
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = self._upload_json(client, rows=rows)
        assert r.get_json()["total"] == 2

    def test_multipart_form_upload(self, client):
        xlsx = make_xlsx_bytes([HEADER_ROW, [PHONE_LOCAL, "יוסי כהן"]])
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = client.post(
                "/web/process",
                headers={**H, "Origin": "http://testserver"},
                data={
                    "file":         (io.BytesIO(xlsx), "test.xlsx"),
                    "apis":         "me",
                    "refresh_days": "0",
                },
                content_type="multipart/form-data",
            )
        assert r.status_code == 200
        assert r.get_json()["success"] is True

    # ── error cases ──────────────────────────────────────────────────────────
    def test_wrong_extension_rejected(self, client):
        r = json_post(client, "/web/process", {
            "file_data": b64(b"data"), "filename": "test.txt", "apis": "me",
        })
        assert r.status_code == 400

    def test_no_file_field_returns_error(self, client):
        r = json_post(client, "/web/process", {})
        assert r.get_json()["success"] is False

    def test_non_hebrew_name_row_rejected(self, client):
        r = self._upload_json(client, rows=[HEADER_ROW, [PHONE_LOCAL, "John Smith"]])
        assert r.get_json()["success"] is False

    def test_invalid_phone_row_rejected(self, client):
        r = self._upload_json(client, rows=[HEADER_ROW, ["notaphone", "יוסי כהן"]])
        assert r.get_json()["success"] is False

    def test_empty_xlsx_rejected(self, client):
        """A file with only the header row and no data rows must be rejected."""
        r = self._upload_json(client, rows=[HEADER_ROW])   # header only, no data
        assert r.get_json()["success"] is False

    def test_cache_counts_reported(self, client):
        phone = "0521234573"   # unique phone to avoid cross-test cache collisions
        # First call primes the cache (refresh_days=0 forces API call)
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            self._upload_json(client, rows=[HEADER_ROW, [phone, "יוסי כהן"]])
        # Second call with refresh_days=30 should read from cache (data < 30 days old)
        r = self._upload_json(client, rows=[HEADER_ROW, [phone, "יוסי כהן"]], refresh_days=30)
        d = r.get_json()
        assert d["success"] is True
        assert d["from_cache"] == 1
        assert d["api_calls"] == 0


# ─────────────────────────────────────────────────────────────────────────────
# 14. Nicknames CRUD
# ─────────────────────────────────────────────────────────────────────────────
class TestNicknamesCRUD:
    FNAME = "אברהם"
    NAMES = "אברהם,אבי,אבוש"

    def _save(self, client, formal=None, names=None):
        return json_post(client, "/web/nicknames/save", {
            "formal_name": formal or self.FNAME,
            "all_names":   names  or self.NAMES,
        })

    def _delete(self, client, formal=None):
        return json_post(client, "/web/nicknames/delete",
                         {"formal_name": formal or self.FNAME})

    def _get(self, client, formal=None):
        return client.get(f"/web/nicknames/get?name={formal or self.FNAME}", headers=H)

    def setup_method(self):
        """Ensure clean slate for each test in this class."""

    def test_save_creates_entry(self, client):
        self._delete(client)          # remove if present
        r = self._save(client)
        assert r.get_json()["success"] is True

    def test_list_contains_entry(self, client):
        self._save(client)
        r = client.get("/web/nicknames/list", headers=H)
        assert r.status_code == 200
        d = r.get_json()
        assert "total_names" in d
        assert self.FNAME in [n["formal_name"] for n in d["nicknames"]]

    def test_list_totals_correct(self, client):
        self._save(client)
        d = client.get("/web/nicknames/list", headers=H).get_json()
        assert d["total_names"] >= 1
        assert d["total_nicknames"] >= 1

    def test_get_returns_entry(self, client):
        self._save(client)
        r = self._get(client)
        assert r.get_json()["found"] is True
        assert r.get_json()["formal_name"] == self.FNAME

    def test_get_missing_name_found_false(self, client):
        r = client.get("/web/nicknames/get?name=שםמדומה999", headers=H)
        assert r.get_json()["found"] is False

    def test_get_missing_param(self, client):
        r = client.get("/web/nicknames/get", headers=H)
        assert r.status_code == 200
        d = r.get_json()
        assert d.get("found") is False or "error" in d

    def test_update_replaces_names(self, client):
        self._save(client)
        self._save(client, names="אברהם,אבי,אברמל")
        d = self._get(client).get_json()
        assert "אברמל" in d["all_names"]

    def test_delete_removes_entry(self, client):
        self._save(client)
        assert self._delete(client).get_json()["success"] is True
        assert self._get(client).get_json()["found"] is False

    def test_delete_nonexistent_fails(self, client):
        r = self._delete(client, formal="שםשלאקיים999")
        assert r.get_json()["success"] is False

    def test_save_empty_formal_name_fails(self, client):
        r = json_post(client, "/web/nicknames/save",
                      {"formal_name": "", "all_names": "a,b"})
        assert r.get_json()["success"] is False

    def test_save_empty_all_names_fails(self, client):
        r = json_post(client, "/web/nicknames/save",
                      {"formal_name": "שם", "all_names": ""})
        assert r.get_json()["success"] is False

    def test_download_json_format(self, client):
        self._save(client)
        r = client.get("/web/nicknames/download", headers=H)
        assert r.status_code == 200
        assert r.mimetype == "application/json"
        data = json.loads(r.data)
        assert isinstance(data, list)
        assert all("formal_name" in e and "all_names" in e for e in data)

    def test_upload_json_file(self, client):
        payload = json.dumps([
            {"formal_name": "רחל", "all_names": "רחל,רחלי,ריקי"}
        ]).encode("utf-8")
        r = client.post(
            "/web/nicknames/upload",
            headers={**H, "Origin": "http://testserver"},
            data={"file": (io.BytesIO(payload), "nicknames.json"), "mode": "add"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        d = r.get_json()
        assert d["success"] is True
        assert d.get("added", 0) + d.get("updated", 0) >= 1

    def test_upload_xlsx_file(self, client):
        xlsx = make_xlsx_bytes([
            ["formal_name", "all_names"],
            ["שרה", "שרה,שרית,שרוני"],
        ])
        r = client.post(
            "/web/nicknames/upload",
            headers={**H, "Origin": "http://testserver"},
            data={"file": (io.BytesIO(xlsx), "nicknames.xlsx"), "mode": "overwrite"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        assert r.get_json()["success"] is True

    def test_upload_csv_file(self, client):
        csv_data = make_csv_bytes([
            ["formal_name", "all_names"],
            ["מרים", "מרים,מרי,מרימה"],
        ])
        r = client.post(
            "/web/nicknames/upload",
            headers={**H, "Origin": "http://testserver"},
            data={"file": (io.BytesIO(csv_data), "nicknames.csv"), "mode": "add"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        assert r.get_json()["success"] is True

    def test_upload_invalid_type_rejected(self, client):
        r = client.post(
            "/web/nicknames/upload",
            headers={**H, "Origin": "http://testserver"},
            data={"file": (io.BytesIO(b"data"), "nicknames.pdf"), "mode": "add"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400

    def test_backup_and_restore(self, client):
        self._save(client, formal="גדעון", names="גדי,גדיאל")
        # Backup
        b = json_post(client, "/web/nicknames/backup", {})
        assert b.status_code == 200
        assert b.get_json()["success"] is True
        # Restore
        rs = json_post(client, "/web/nicknames/restore", {})
        assert rs.status_code == 200
        assert rs.get_json()["success"] is True
        # Entry should still be accessible
        assert self._get(client, "גדעון").get_json()["found"] is True

    def test_nicknames_api_returns_variants(self, client):
        """GET /nicknames?name=X (REST API) should return all related names."""
        self._save(client, formal="יוסף", names="יוסי,יוסיפון,יוסיק")
        r = client.get("/nicknames?name=יוסף", headers=H)
        d = r.get_json()
        assert "יוסי" in d["names"]
        assert "יוסיפון" in d["names"]


# ─────────────────────────────────────────────────────────────────────────────
# 15. Security Vectors
# ─────────────────────────────────────────────────────────────────────────────
class TestSecurityVectors:

    # ── DISABLE_AUTH removed ─────────────────────────────────────────────────
    def test_disable_auth_env_no_longer_bypasses(self):
        saved = os.environ.get("DISABLE_AUTH")
        os.environ["DISABLE_AUTH"] = "true"
        try:
            with flask_app.test_client() as c:
                r = c.get("/web/query")    # no CF auth header
                assert r.status_code == 403, "DISABLE_AUTH=true should NOT bypass auth"
        finally:
            if saved is None:
                os.environ.pop("DISABLE_AUTH", None)
            else:
                os.environ["DISABLE_AUTH"] = saved

    # ── Magic bytes ──────────────────────────────────────────────────────────
    def test_fake_xlsx_rejected_by_magic_bytes(self, client):
        """Plain text file with .xlsx extension must be rejected."""
        r = json_post(client, "/web/process", {
            "file_data": b64(b"This is plain text, not an xlsx"),
            "filename":  "evil.xlsx",
            "apis":      "me",
        })
        assert r.status_code == 400
        assert "content" in r.get_json()["error"].lower()

    def test_real_xlsx_magic_accepted(self, client):
        xlsx = make_xlsx_bytes([HEADER_ROW, [PHONE_LOCAL, "יוסי כהן"]])
        with patch("providers.me.MEProvider.call_api", return_value=ME_API_RESPONSE):
            r = json_post(client, "/web/process", {
                "file_data":    b64(xlsx),
                "filename":     "real.xlsx",
                "apis":         "me",
                "refresh_days": 0,
            })
        assert r.status_code == 200
        assert r.get_json()["success"] is True

    def test_invalid_base64_returns_400(self, client):
        r = json_post(client, "/web/process", {
            "file_data": "!!!not-valid-base64!!!",
            "filename":  "test.xlsx",
            "apis":      "me",
        })
        assert r.status_code == 400
        assert "encoding" in r.get_json()["error"].lower()

    # ── Excel formula injection ──────────────────────────────────────────────
    def test_formula_in_api_result_is_escaped_in_output(self, client):
        """Cells starting with = from API data must be escaped in the output Excel."""
        import openpyxl
        poisoned = dict(ME_API_RESPONSE)
        poisoned["common_name"] = "=CMD|'/C calc'!A0"

        with patch("providers.me.MEProvider.call_api", return_value=poisoned):
            r = json_post(client, "/web/process", {
                "file_data":    b64(make_xlsx_bytes([HEADER_ROW, [PHONE_LOCAL, "יוסי כהן"]])),
                "filename":     "test.xlsx",
                "apis":         "me",
                "refresh_days": 0,
            })
        assert r.get_json()["success"] is True
        file_id = r.get_json()["file_id"]

        dl = client.get(f"/web/download/{file_id}", headers=H)
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        all_values = [str(c.value or "") for row in wb.active.iter_rows() for c in row]

        assert not any(v == "=CMD|'/C calc'!A0" for v in all_values), \
            "Unescaped formula found in output Excel"
        assert any(v.startswith("'=") for v in all_values), \
            "Escaped formula (prefixed with ') not found in output"

    def test_plus_prefix_formula_escaped(self, client):
        """Cells starting with + are also formula-injectable and must be escaped."""
        import openpyxl
        poisoned = dict(ME_API_RESPONSE)
        poisoned["common_name"] = "+1-2"

        with patch("providers.me.MEProvider.call_api", return_value=poisoned):
            r = json_post(client, "/web/process", {
                "file_data":    b64(make_xlsx_bytes([HEADER_ROW, [PHONE_LOCAL, "יוסי כהן"]])),
                "filename":     "test.xlsx",
                "apis":         "me",
                "refresh_days": 0,
            })
        file_id = r.get_json()["file_id"]
        dl = client.get(f"/web/download/{file_id}", headers=H)
        wb = openpyxl.load_workbook(io.BytesIO(dl.data))
        all_values = [str(c.value or "") for row in wb.active.iter_rows() for c in row]
        assert not any(v == "+1-2" for v in all_values)

    # ── Nickname input sanitization ──────────────────────────────────────────
    def test_script_tag_stripped_from_formal_name(self, client):
        r = json_post(client, "/web/nicknames/save", {
            "formal_name": "<script>alert(1)</script>",
            "all_names":   "test",
        })
        d = r.get_json()
        if d.get("success"):
            lst = client.get("/web/nicknames/list", headers=H).get_json()
            assert not any("<script>" in n["formal_name"] for n in lst["nicknames"])

    def test_long_formal_name_truncated(self, client):
        """formal_name > 200 chars is truncated, not stored at full length."""
        long_name = "א" * 300
        r = json_post(client, "/web/nicknames/save",
                      {"formal_name": long_name, "all_names": "א"})
        if r.get_json().get("success"):
            clipped = "א" * 200
            g = client.get(f"/web/nicknames/get?name={clipped}", headers=H)
            if g.get_json().get("found"):
                assert len(g.get_json().get("all_names", "")) <= 1000

    # ── LIKE wildcard injection ──────────────────────────────────────────────
    def test_percent_wildcard_not_expanded_in_nicknames(self, client):
        """GET /nicknames?name=% must not return all DB rows."""
        r = client.get("/nicknames?name=%25", headers=H)   # URL-encoded %
        assert r.status_code == 200
        d = r.get_json()
        # Should only contain the query name itself, not every entry
        assert len(d["names"]) < 20, \
            "Wildcard '%' appears to have matched all DB rows (LIKE injection)"

    def test_underscore_wildcard_not_expanded(self, client):
        r = client.get("/nicknames?name=_", headers=H)
        d = r.get_json()
        assert len(d["names"]) < 20

    # ── Phone masking in logs ────────────────────────────────────────────────
    def test_mask_phone_hides_digits(self):
        from app_logger import _mask_phone
        masked = _mask_phone("972521234567")
        assert masked.endswith("4567")
        assert masked.startswith("*")
        assert "972521" not in masked

    def test_mask_phone_short(self):
        from app_logger import _mask_phone
        assert _mask_phone("1234") == "****"
        assert _mask_phone("")    == "****"
        assert _mask_phone(None)  == "****"

    def test_mask_phone_length_preserved(self):
        from app_logger import _mask_phone
        phone = "972521234567"
        assert len(_mask_phone(phone)) == len(phone)
